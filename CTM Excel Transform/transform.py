import openpyxl
import os

def transform(input_file,output_file):
    if input_file == "":
        return 0
    master_file = output_file
    # master_file = None

    # open our master spreadsheet or create one if none was provided
    if master_file!="":
        try:
            master = openpyxl.load_workbook(master_file)
        except Exception as e:
            print("Spreadsheet does not exist. Creating a new one.")
            master_file=""
    print("starting transform")
    if master_file=="":
        master = openpyxl.Workbook()
        master_file = os.path.join(os.path.dirname(input_file),"output.xlsx")

    sheet_master = master.get_sheet_by_name(master.get_sheet_names()[0])
    sheet_master.title = "Master"

    print("created/opened spreadsheet")

    # determine the row we start writing data
    master_row = sheet_master.max_row + 1 

    # get the data we're importing into master
    raw = openpyxl.load_workbook(input_file)
    sheet_raw = raw.get_sheet_by_name(raw.get_sheet_names()[0])

    print("loaded import sheet")
    col = 1
    #gather column data
    while sheet_raw.cell(row=1,column=col).value!="School Name":
        col += 1
    start = col

    while sheet_raw.cell(row=1,column=col).value!="Unique ID":
        col += 1
    end = col

    # define special columns for additional handling
    special_questions = ['Sensory Kit (Includes noise dampening headphones, fidgets, sunglasses)','Assistive Listening Devices (Overture ONLY)','Wheelchair Accessible Seating','ASL Interpretation','Specially reserved seating for students with visual impairments']
    grade_question = 'What grade(s) are you bringing?'
    grades = ['4k','Kindergarten','1st Grade','2nd Grade','3rd Grade','4th Grade','5th Grade','6th Grade','7th Grade','8th Grade','9th Grade','10th Grade','11th Grade','12th Grade']

    # if no data is in master, create column headers
    if master_row==2:
        sheet_master.cell(row=1,column=1).value = "Time"
        sheet_master.cell(row=1,column=2).value = "Show"
        sheet_master.cell(row=1,column=3).value = "Quantity"
        master_col = 4
        for col in range(start,end+1):
            # create a column for each possible grade
            if sheet_raw.cell(row=1,column=col).value==grade_question:
                for i in range(len(grades)):
                    sheet_master.cell(row=1,column=master_col).value = grades[i]
                    master_col += 1        
            else:
                sheet_master.cell(row=1,column=master_col).value = sheet_raw.cell(row=1,column=col).value
                master_col += 1
    # loop over all rows
    for row in range(2,sheet_raw.max_row+1):
        # loop over all shows
        for col in range(2,start):
            if sheet_raw.cell(row=row,column=col).value!=None:
                # get ticket information   
                ticket_data = sheet_raw.cell(row=row,column=col).value
                ticket_data = [i.split(" = ") for i in ticket_data.split("\n")]
                ticket_data = dict(ticket_data)

                # write the time, show, and quantity to master spreadsheet
                sheet_master.cell(row=master_row,column=1).value = sheet_raw.cell(row=row,column=1).value
                sheet_master.cell(row=master_row,column=2).value = sheet_raw.cell(row=1,column=col).value
                sheet_master.cell(row=master_row,column=3).value = int(ticket_data['quantity'])
                master_col = 4
                # write the meta data of order
                for col in range(start,end+1):
                    cell_value = sheet_raw.cell(row=row,column=col).value
                    # handle null values
                    if cell_value is None:
                        master_col += 1
                    # special handling of complex questions
                    elif sheet_raw.cell(row=1,column=col).value in special_questions:
                        question_data = cell_value
                        question_data = [i.split(" = ") for i in question_data.split("\n")]
                        question_data = dict(question_data)
                        sheet_master.cell(row=master_row,column=master_col).value = int(question_data['quantity'])
                        master_col += 1
                    # handle grade level question specially
                    elif sheet_raw.cell(row=1,column=col).value == grade_question:
                        question_data = cell_value
                        question_data = question_data.split("\n")
                        for i in range(len(grades)):
                            if grades[i] in question_data:
                                sheet_master.cell(row=master_row,column=master_col).value = 'Yes'
                            #else: 
                            #    sheet_master.cell(row=master_row,column=master_col).value = 'No'
                            master_col += 1        
                    else:
                        sheet_master.cell(row=master_row,column=master_col).value = cell_value
                        master_col += 1
                master_row += 1
    print("finished parsing")
    master.save(master_file)
    print("saved")
